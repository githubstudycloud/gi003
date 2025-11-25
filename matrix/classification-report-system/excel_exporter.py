"""
Excel报表导出器
支持混淆矩阵、详细数据、多sheet导出
"""
from typing import List, Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from data_model import ClassificationRecord, DataRepository, FilterCriteria
from confusion_matrix import ConfusionMatrixGenerator
import numpy as np


class ExcelExporter:
    """Excel导出器"""

    def __init__(self, repository: DataRepository):
        self.repository = repository

    def export_full_report(
        self,
        output_path: str,
        filter_criteria: FilterCriteria = None
    ):
        """
        导出完整报告到Excel
        包含多个sheet:
        1. 汇总统计
        2. 总体混淆矩阵
        3. 各一级分类混淆矩阵
        4. 详细数据列表
        """
        # 获取记录
        if filter_criteria:
            records = self.repository.filter_records(filter_criteria)
        else:
            records = self.repository.get_all_records()

        if not records:
            raise ValueError("没有找到匹配的记录")

        # 生成混淆矩阵数据
        generator = ConfusionMatrixGenerator(records)
        report_data = generator.generate_detailed_report()

        # 创建Excel工作簿
        wb = Workbook()
        wb.remove(wb.active)  # 删除默认sheet

        # 1. 汇总统计sheet
        self._create_summary_sheet(wb, report_data["summary"])

        # 2. 总体混淆矩阵sheet
        self._create_confusion_matrix_sheet(
            wb,
            "总体混淆矩阵",
            report_data["overall"]
        )

        # 3. 各一级分类混淆矩阵sheet
        for category, matrix_data in report_data["by_primary_category"].items():
            sheet_name = f"分类-{category[:20]}"  # 限制sheet名称长度
            self._create_confusion_matrix_sheet(wb, sheet_name, matrix_data)

        # 4. 详细数据列表sheet
        self._create_detail_data_sheet(wb, records)

        # 保存文件
        wb.save(output_path)
        return output_path

    def _create_summary_sheet(self, wb: Workbook, summary: Dict):
        """创建汇总统计sheet"""
        ws = wb.create_sheet("汇总统计")

        # 标题
        ws['A1'] = "汇总统计报告"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:B1')

        # 基本统计
        row = 3
        ws[f'A{row}'] = "指标"
        ws[f'B{row}'] = "数值"
        self._style_header_row(ws, row)

        stats = [
            ("总记录数", summary['total_records']),
            ("通过数 (PASS)", summary['passed']),
            ("失败数 (FAIL)", summary['failed']),
            ("准确率 (%)", summary['accuracy']),
        ]

        row += 1
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

        # 维度统计
        row += 1
        ws[f'A{row}'] = "维度统计"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        ws[f'A{row}'] = "维度"
        ws[f'B{row}'] = "唯一值数量"
        self._style_header_row(ws, row)

        row += 1
        for key, value in summary['unique_counts'].items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            row += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15

    def _create_confusion_matrix_sheet(
        self,
        wb: Workbook,
        sheet_name: str,
        matrix_data: Dict
    ):
        """创建混淆矩阵sheet"""
        ws = wb.create_sheet(sheet_name)

        matrix = np.array(matrix_data["matrix"])
        precision = matrix_data["precision"]
        recall = matrix_data["recall"]
        total_expected = matrix_data["total_expected"]
        total_actual = matrix_data["total_actual"]

        # 标题
        ws['A1'] = sheet_name
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells(f'A1:{get_column_letter(19)}1')

        # 表头 (第3行)
        row = 3
        ws[f'A{row}'] = "实际\\预测"

        # 预测值列 (0-15)
        for i in range(16):
            col = get_column_letter(i + 2)
            ws[f'{col}{row}'] = f"预测{i}"

        ws[f'{get_column_letter(18)}{row}'] = "SUM"
        ws[f'{get_column_letter(19)}{row}'] = "召回率(%)"

        self._style_header_row(ws, row, 19)

        # 数据行 (实际0-15)
        row = 4
        for i in range(16):
            ws[f'A{row}'] = f"实际{i}"

            # 混淆矩阵数据
            for j in range(16):
                col = get_column_letter(j + 2)
                ws[f'{col}{row}'] = matrix[i][j]

                # 对角线高亮（正确预测）
                if i == j and matrix[i][j] > 0:
                    ws[f'{col}{row}'].fill = PatternFill(
                        start_color="90EE90",
                        end_color="90EE90",
                        fill_type="solid"
                    )

            # SUM列
            ws[f'{get_column_letter(18)}{row}'] = total_actual[i]

            # 召回率列
            ws[f'{get_column_letter(19)}{row}'] = recall[i]

            row += 1

        # SUM行
        ws[f'A{row}'] = "SUM"
        ws[f'A{row}'].font = Font(bold=True)

        for i in range(16):
            col = get_column_letter(i + 2)
            ws[f'{col}{row}'] = total_expected[i]

        ws[f'{get_column_letter(18)}{row}'] = sum(total_expected)
        ws[f'{get_column_letter(19)}{row}'] = "-"

        row += 1

        # 精准率行
        ws[f'A{row}'] = "精准率(%)"
        ws[f'A{row}'].font = Font(bold=True)

        for i in range(16):
            col = get_column_letter(i + 2)
            ws[f'{col}{row}'] = precision[i]

        ws[f'{get_column_letter(18)}{row}'] = "-"
        ws[f'{get_column_letter(19)}{row}'] = "-"

        # 调整列宽
        ws.column_dimensions['A'].width = 12
        for i in range(2, 20):
            ws.column_dimensions[get_column_letter(i)].width = 8

    def _create_detail_data_sheet(self, wb: Workbook, records: List[ClassificationRecord]):
        """创建详细数据列表sheet"""
        ws = wb.create_sheet("详细数据")

        # 表头
        headers = [
            "一级分类", "二级分类", "预期值", "实际值", "状态(Pass/Fail)",
            "用例", "场景", "垂类", "因子", "因子值",
            "测试ID", "时间戳", "备注"
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="4472C4",
                end_color="4472C4",
                fill_type="solid"
            )
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')

        # 数据行
        for row_idx, record in enumerate(records, 2):
            data = [
                record.primary_category,
                record.secondary_category,
                record.expected_value,
                record.actual_value,
                record.status.value.upper(),
                record.use_case,
                record.scenario,
                record.vertical,
                record.factor,
                record.factor_value,
                record.test_id or "",
                record.timestamp or "",
                record.notes or ""
            ]

            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value

                # PASS/FAIL着色
                if col_idx == 5:  # 状态列
                    if value == "PASS":
                        cell.fill = PatternFill(
                            start_color="90EE90",
                            end_color="90EE90",
                            fill_type="solid"
                        )
                    else:
                        cell.fill = PatternFill(
                            start_color="FFB6C1",
                            end_color="FFB6C1",
                            fill_type="solid"
                        )

        # 调整列宽
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        # 冻结首行
        ws.freeze_panes = "A2"

    def _style_header_row(self, ws, row: int, num_cols: int = 2):
        """设置表头行样式"""
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4472C4",
                end_color="4472C4",
                fill_type="solid"
            )
            cell.alignment = Alignment(horizontal='center')

    def export_simple_excel(
        self,
        output_path: str,
        records: List[ClassificationRecord] = None
    ):
        """
        导出简单的Excel（仅包含详细数据）
        使用pandas快速导出
        """
        if records is None:
            records = self.repository.get_all_records()

        # 转换为DataFrame
        data = [record.to_dict() for record in records]
        df = pd.DataFrame(data)

        # 重命名列
        df.columns = [
            "一级分类", "二级分类", "预期值", "实际值", "状态",
            "用例", "场景", "垂类", "因子", "因子值",
            "时间戳", "测试ID", "备注"
        ]

        # 导出到Excel
        df.to_excel(output_path, index=False, sheet_name="数据")
        return output_path
